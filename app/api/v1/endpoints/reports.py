import csv
import io
from datetime import datetime, timedelta, timezone

from fastapi import APIRouter, Depends, HTTPException, Query
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session

from app.db.session import get_db
from app.core.auth import get_current_active_user
from app.models.appointment import Appointment, AppointmentStatus
from app.models.slot import Slot
from app.models.client import Client
from app.models.provider import Provider
from app.models.service import Service
from app.models.user import UserRole

router = APIRouter()

TZ_TBILISI = timezone(timedelta(hours=4))

STATUS_KA = {
    "pending":   "მოლოდინში",
    "confirmed": "დადასტურებული",
    "cancelled": "გაუქმებული",
    "completed": "დასრულებული",
    "no_show":   "არ გამოცხადდა",
}


def _to_tbilisi(dt: datetime | None) -> str:
    if not dt:
        return ""
    if dt.tzinfo is None:
        dt = dt.replace(tzinfo=timezone.utc)
    return dt.astimezone(TZ_TBILISI).strftime("%Y-%m-%d %H:%M")


def _base_query(
    db: Session, current_user, tenant_id: str,
    date_from: str | None, date_to: str | None,
    provider_id: str | None, service_id: str | None,
    status: AppointmentStatus | None,
):
    q = (
        db.query(Appointment, Slot, Client, Provider, Service)
        .join(Slot, Slot.id == Appointment.slot_id)
        .join(Client, Client.id == Appointment.client_id)
        .join(Provider, Provider.id == Slot.provider_id)
        .outerjoin(Service, Service.id == Slot.service_id)
        .filter(Appointment.tenant_id == tenant_id)
    )

    if current_user.role == UserRole.provider and current_user.provider_id:
        q = q.filter(Slot.provider_id == current_user.provider_id)
    elif provider_id:
        q = q.filter(Slot.provider_id == provider_id)

    if service_id:
        q = q.filter(Slot.service_id == service_id)
    if status:
        q = q.filter(Appointment.status == status)

    if date_from:
        try:
            df = datetime.strptime(date_from, "%Y-%m-%d")
            q = q.filter(Slot.starts_at >= df)
        except ValueError:
            raise HTTPException(400, "date_from ფორმატი: YYYY-MM-DD")
    if date_to:
        try:
            dt = datetime.strptime(date_to, "%Y-%m-%d") + timedelta(days=1)
            q = q.filter(Slot.starts_at < dt)
        except ValueError:
            raise HTTPException(400, "date_to ფორმატი: YYYY-MM-DD")

    return q.order_by(Slot.starts_at.desc())


def _row(a: Appointment, s: Slot, c: Client, p: Provider, svc: Service | None) -> dict:
    return {
        "id":            a.id,
        "date":          s.starts_at.strftime("%Y-%m-%d") if s.starts_at else "",
        "time":          s.starts_at.strftime("%H:%M") if s.starts_at else "",
        "client_name":   f"{c.first_name} {c.last_name}",
        "client_phone":  c.phone or "",
        "personal_id":   c.personal_id or "",
        "provider_name": f"{p.first_name} {p.last_name}",
        "service_name":  (svc.name_ka if svc else "") or "",
        "status":        a.status.value if a.status else "",
        "status_ka":     STATUS_KA.get(a.status.value, a.status.value) if a.status else "",
        "notes":         a.notes or "",
        "created_at":    _to_tbilisi(a.created_at),
        "cancelled_by":  a.cancelled_by or "",
        "cancelled_at":  _to_tbilisi(a.cancelled_at),
        "last_modified_by": a.last_modified_by or "",
        "modified_from_ip": a.modified_from_ip or "",
        "modified_from_ua": a.modified_from_ua or "",
    }


@router.get("/")
def report_list(
    tenant_id:   str = Query(...),
    date_from:   str | None = Query(None),
    date_to:     str | None = Query(None),
    provider_id: str | None = Query(None),
    service_id:  str | None = Query(None),
    status:      AppointmentStatus | None = Query(None),
    limit:       int = Query(200, le=1000),
    offset:      int = Query(0, ge=0),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
):
    q = _base_query(db, current_user, tenant_id, date_from, date_to,
                    provider_id, service_id, status)

    total = q.count()
    rows = [_row(*r) for r in q.limit(limit).offset(offset).all()]

    sq = _base_query(db, current_user, tenant_id, date_from, date_to,
                     provider_id, service_id, None)
    counts = {st.value: 0 for st in AppointmentStatus}
    for a, *_ in sq.all():
        counts[a.status.value] = counts.get(a.status.value, 0) + 1

    return {
        "total": total,
        "rows": rows,
        "summary": {
            "all": sum(counts.values()),
            **counts,
        },
    }


EXPORT_HEADERS = [
    ("date",          "თარიღი"),
    ("time",          "დრო"),
    ("client_name",   "კლიენტი"),
    ("client_phone",  "ტელეფონი"),
    ("personal_id",   "პირადი ნომერი"),
    ("provider_name", "ექიმი"),
    ("service_name",  "სერვისი"),
    ("status_ka",     "სტატუსი"),
    ("cancelled_by",  "გააუქმა"),
    ("cancelled_at",  "გაუქმების დრო"),
    ("last_modified_by", "ბოლოს შეცვალა"),
    ("modified_from_ip", "IP მისამართი"),
    ("modified_from_ua", "სისტემა"),
    ("notes",         "შენიშვნა"),
    ("created_at",    "შეიქმნა"),
]


@router.get("/export")
def report_export(
    tenant_id:   str = Query(...),
    format:      str = Query("csv", pattern="^(csv|xlsx)$"),
    date_from:   str | None = Query(None),
    date_to:     str | None = Query(None),
    provider_id: str | None = Query(None),
    service_id:  str | None = Query(None),
    status:      AppointmentStatus | None = Query(None),
    db: Session = Depends(get_db),
    current_user = Depends(get_current_active_user),
):
    q = _base_query(db, current_user, tenant_id, date_from, date_to,
                    provider_id, service_id, status)
    rows = [_row(*r) for r in q.all()]

    stamp = datetime.now().strftime("%Y%m%d_%H%M")

    if format == "csv":
        buf = io.StringIO()
        buf.write("\ufeff")
        w = csv.writer(buf, delimiter=";")
        w.writerow([h for _, h in EXPORT_HEADERS])
        for r in rows:
            w.writerow([r[k] for k, _ in EXPORT_HEADERS])
        buf.seek(0)
        return StreamingResponse(
            iter([buf.getvalue()]),
            media_type="text/csv; charset=utf-8",
            headers={"Content-Disposition": f"attachment; filename=report_{stamp}.csv"},
        )

    try:
        from openpyxl import Workbook
        from openpyxl.styles import Font, PatternFill
        from openpyxl.utils import get_column_letter
    except ImportError:
        raise HTTPException(501, "openpyxl არ არის დაყენებული")

    wb = Workbook()
    ws = wb.active
    ws.title = "ანგარიში"

    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="1D9E75", end_color="1D9E75", fill_type="solid")

    for col, (_, h) in enumerate(EXPORT_HEADERS, start=1):
        cell = ws.cell(row=1, column=col, value=h)
        cell.font = header_font
        cell.fill = header_fill

    for i, r in enumerate(rows, start=2):
        for col, (k, _) in enumerate(EXPORT_HEADERS, start=1):
            ws.cell(row=i, column=col, value=r[k])

    for col, (k, h) in enumerate(EXPORT_HEADERS, start=1):
        max_len = max([len(str(h))] + [len(str(r[k])) for r in rows[:500]] or [10])
        ws.column_dimensions[get_column_letter(col)].width = min(max_len + 3, 40)

    ws.freeze_panes = "A2"

    out = io.BytesIO()
    wb.save(out)
    out.seek(0)
    return StreamingResponse(
        out,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=report_{stamp}.xlsx"},
    )
