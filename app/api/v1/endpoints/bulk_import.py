"""
Bulk Import — Excel-იდან პროვაიდერების და სერვისების იმპორტი.
"""

from fastapi import APIRouter, UploadFile, File, Depends, HTTPException
from sqlalchemy.orm import Session
from io import BytesIO
import openpyxl

from app.db.session import get_db
from app.core.auth import get_current_user, require_admin
from app.models.provider import Provider, ProviderService
from app.models.service import Service
from app.models.user import User

router = APIRouter(prefix="/admin/import", tags=["bulk-import"])


def _parse_bool(val) -> bool:
    if val is None or val == "":
        return True
    if isinstance(val, bool):
        return val
    return str(val).strip().upper() in ("TRUE", "1", "YES")


def _clean(val) -> str:
    if val is None:
        return ""
    return str(val).strip()


@router.post("/bulk-upload")
async def bulk_upload(
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    current_user: User = Depends(require_admin),
):
    tenant_id = current_user.tenant_id
    if not tenant_id:
        raise HTTPException(400, "Platform admin-ს tenant არ აქვს მინიჭებული")

    if not file.filename.endswith((".xlsx", ".xls")):
        raise HTTPException(400, "მხოლოდ .xlsx ფაილი მიიღება")

    content = await file.read()
    try:
        wb = openpyxl.load_workbook(BytesIO(content), read_only=True, data_only=True)
    except Exception as e:
        raise HTTPException(400, f"ფაილის წაკითხვა ვერ მოხერხდა: {str(e)}")

    results = {
        "providers": {"created": 0, "updated": 0, "errors": []},
        "services": {"created": 0, "updated": 0, "errors": []},
        "mappings": {"created": 0, "skipped": 0, "errors": []},
    }

    svc_sheet_name = None
    for name in wb.sheetnames:
        if "სერვის" in name.lower() or "service" in name.lower():
            svc_sheet_name = name
            break

    if svc_sheet_name:
        ws = wb[svc_sheet_name]
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        for row_idx, row in enumerate(rows, start=3):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            try:
                code = _clean(row[0])
                name_ka = _clean(row[1])
                name_en = _clean(row[2]) if len(row) > 2 else ""
                duration = int(row[3]) if len(row) > 3 and row[3] else 30
                color = _clean(row[4]) if len(row) > 4 and row[4] else "#1D9E75"
                active = _parse_bool(row[5] if len(row) > 5 else None)
                if not code or not name_ka:
                    results["services"]["errors"].append(f"რიგი {row_idx}: code და name_ka სავალდებულოა")
                    continue
                existing = db.query(Service).filter(Service.tenant_id == tenant_id, Service.code == code).first()
                if existing:
                    existing.name_ka = name_ka
                    existing.name_en = name_en or existing.name_en
                    existing.duration_min = duration
                    existing.color = color
                    existing.active = active
                    results["services"]["updated"] += 1
                else:
                    db.add(Service(tenant_id=tenant_id, code=code, name_ka=name_ka, name_en=name_en, duration_min=duration, color=color, active=active))
                    results["services"]["created"] += 1
            except Exception as e:
                results["services"]["errors"].append(f"რიგი {row_idx}: {str(e)}")
        db.flush()

    prov_sheet_name = None
    for name in wb.sheetnames:
        if "პროვაიდერ" in name.lower() or "provider" in name.lower():
            prov_sheet_name = name
            break

    if prov_sheet_name:
        ws = wb[prov_sheet_name]
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        for row_idx, row in enumerate(rows, start=3):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            try:
                first_name = _clean(row[0])
                last_name = _clean(row[1])
                specialty = _clean(row[2]) if len(row) > 2 else ""
                phone = _clean(row[3]) if len(row) > 3 else ""
                email = _clean(row[4]) if len(row) > 4 else ""
                active = _parse_bool(row[5] if len(row) > 5 else None)
                if not first_name or not last_name:
                    results["providers"]["errors"].append(f"რიგი {row_idx}: სახელი და გვარი სავალდებულოა")
                    continue
                existing = None
                if email:
                    existing = db.query(Provider).filter(Provider.tenant_id == tenant_id, Provider.email == email).first()
                if existing:
                    existing.first_name = first_name
                    existing.last_name = last_name
                    existing.specialty = specialty or existing.specialty
                    existing.phone = phone or existing.phone
                    existing.active = active
                    results["providers"]["updated"] += 1
                else:
                    db.add(Provider(tenant_id=tenant_id, first_name=first_name, last_name=last_name, specialty=specialty, phone=phone, email=email, active=active))
                    results["providers"]["created"] += 1
            except Exception as e:
                results["providers"]["errors"].append(f"რიგი {row_idx}: {str(e)}")
        db.flush()

    map_sheet_name = None
    for name in wb.sheetnames:
        if "↔" in name or "mapping" in name.lower():
            map_sheet_name = name
            break

    if map_sheet_name:
        ws = wb[map_sheet_name]
        rows = list(ws.iter_rows(min_row=3, values_only=True))
        for row_idx, row in enumerate(rows, start=3):
            if not row or all(c is None or str(c).strip() == "" for c in row):
                continue
            try:
                prov_email = _clean(row[0])
                svc_code = _clean(row[1])
                if not prov_email or not svc_code:
                    continue
                if prov_email.startswith("შენიშვნა") or prov_email.startswith("Note"):
                    continue
                provider = db.query(Provider).filter(Provider.tenant_id == tenant_id, Provider.email == prov_email).first()
                if not provider:
                    results["mappings"]["errors"].append(f"რიგი {row_idx}: პროვაიდერი '{prov_email}' ვერ მოიძებნა")
                    continue
                service = db.query(Service).filter(Service.tenant_id == tenant_id, Service.code == svc_code).first()
                if not service:
                    results["mappings"]["errors"].append(f"რიგი {row_idx}: სერვისი '{svc_code}' ვერ მოიძებნა")
                    continue
                existing_map = db.query(ProviderService).filter(ProviderService.provider_id == provider.id, ProviderService.service_id == service.id).first()
                if existing_map:
                    results["mappings"]["skipped"] += 1
                else:
                    db.add(ProviderService(provider_id=provider.id, service_id=service.id))
                    results["mappings"]["created"] += 1
            except Exception as e:
                results["mappings"]["errors"].append(f"რიგი {row_idx}: {str(e)}")

    db.commit()
    wb.close()
    return {"success": True, "results": results}


@router.get("/template-info")
async def template_info(current_user: User = Depends(get_current_user)):
    return {"download_url": "/static/pacsflow_import_template.xlsx"}
